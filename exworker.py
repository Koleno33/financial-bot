from openpyxl import Workbook
from openpyxl.styles import Color, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from database import Record, Currency, Session, engine
from datetime import datetime

class ExcelWorker:
    def __init__(self, records: list[Record], years: list[str], section_currencies: dict[str: list]):
        self.wb = Workbook()
        current_ws = self.wb.active
        current_ws.title = 'Записи'
        self.wb.create_sheet('Промежуточный')
        self.wb.create_sheet('Сводка')
        self.add_records(records)
        self.add_summary(len(records), years, section_currencies)
        self.add_inter(len(records))

    def get_bytes(self):
        res = BytesIO()
        self.wb.save(res)
        return res

    def add_records(self, records: list[Record]):
        col_size = 15
        ws = self.wb['Записи']
        ws.append(["№ Записи", "Раздел", "Валюта", "Значение", "Дата", "Дата добавления", "Комментарий",
                   "Месяц", "Видимое"])
        ws.column_dimensions["A"].width = col_size
        ws.column_dimensions["B"].width = col_size
        ws.column_dimensions["C"].width = col_size
        ws.column_dimensions["D"].width = col_size
        ws.column_dimensions["E"].width = col_size * 2
        ws.column_dimensions["F"].width = col_size * 2
        ws.column_dimensions["G"].width = col_size * 3
        ws.column_dimensions["H"].hidden = True
        ws.column_dimensions["I"].hidden = True

        data = [[r.id, r.section.names[0], r.currency.names[0], r.amount, r.datetime, r.added_datetime,
                 r.comment, r.datetime.month, f"=SUBTOTAL(9,D{i+2})"] # (i+1)+1, т.к. записи начиная со 2 строки
                for i, r in enumerate(records)]
        for row in data:
            ws.append(row)

        ws.auto_filter.ref = f'A1:G{len(data) + 1}'

    def add_inter(self, records_length: int):
        ws = self.wb['Промежуточный']
        ws.sheet_state = "hidden"

        formula = f"""=_xlfn.FILTER(
          Записи!$A$2:$G${records_length + 1},
          (Записи!$E$2:$E${records_length + 1}>=DATE(Сводка!$B$1,$I$2,1)) *
          (Записи!$E$2:$E${records_length + 1}<=DATE(Сводка!$B$1,$I$2 + 1,1)) *
          (IF(Сводка!$B$3="все", TRUE, Записи!$B$2:$B${records_length + 1}=Сводка!$B$3))
        )"""

        match_instruction = '=MATCH(Сводка!B2; {"январь"; "февраль"; "март"; "апрель"; "май"; "июнь"; "июль"; ' \
                            '"август"; "сентябрь"; "октябрь"; "ноябрь"; "декабрь"}; 0)'
        ws["A1"] = ArrayFormula(f"A1:G{records_length}", formula)
        ws["I1"] = ArrayFormula(f"I1:I1", f"=IFERROR(A1=A{records_length},FALSE)")
        ws["I2"] = ArrayFormula(f"I2:I2", match_instruction)
        ws["I3"] = ArrayFormula(f"I3:I3", "=Сводка!B3")

    def add_summary(self, records_length: int, years: list[int], section_currencies):
        ws = self.wb['Сводка']
        months = "январь,февраль,март,апрель,май,июнь,июль,август,сентябрь,октябрь,ноябрь,декабрь"
        col_size = 15

        border = Border(
            left=Side(border_style="thin", color='FF000000'),
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
            diagonal=Side(border_style=None, color='FF000000'),
            diagonal_direction=0,
            outline=Side(border_style=None, color='FF000000'),
            vertical=Side(border_style=None, color='FF000000'),
            horizontal=Side(border_style=None, color='FF000000')
        )

        # Добавляем выбор сети
        sections = list(section_currencies.keys())
        sections_list = ["все"] + sections
        
        ws["A1"] = "Год"
        ws["A2"] = "Месяц"
        ws["A3"] = "Сеть"
        ws.column_dimensions["A"].width = col_size
        ws.column_dimensions["B"].width = col_size
        ws.column_dimensions["C"].width = col_size
        ws.column_dimensions["D"].width = col_size
        ws.column_dimensions["E"].width = col_size
        ws.column_dimensions["F"].width = col_size
        ws.column_dimensions["G"].width = col_size
        ws.column_dimensions["H"].width = col_size
        ws.column_dimensions["I"].width = col_size
        ws.column_dimensions["J"].width = col_size
        ws.column_dimensions["K"].width = col_size
        ws.column_dimensions["L"].width = col_size
        ws.column_dimensions["M"].width = col_size
        ws.column_dimensions["N"].width = col_size
        ws.column_dimensions["G"].hidden = True
        ws.column_dimensions["L"].hidden = True
        ws.column_dimensions["M"].hidden = True
        ws.column_dimensions["N"].hidden = True
        ws.column_dimensions["O"].hidden = True
        
        dv1 = DataValidation(type="list", formula1=f"\"{','.join(years)}\"", allow_blank=False)
        ws.add_data_validation(dv1)
        ws["B1"] = 2025 
        dv1.add(ws["B1"])

        dv2 = DataValidation(type="list", formula1=f"\"{months}\"", allow_blank=False)
        ws.add_data_validation(dv2)
        ws["B2"] = months.split(',')[datetime.now().month - 1]
        dv2.add(ws["B2"])
        ws["B2"].alignment = Alignment(horizontal="right")

        dv3 = DataValidation(type="list", formula1=f'"{",".join(sections_list)}"', allow_blank=False)
        ws.add_data_validation(dv3)
        ws["B3"] = "все"
        dv3.add(ws["B3"])
        ws["B3"].alignment = Alignment(horizontal="right")

        # Добавляем отступ
        ws.append([])
        
        # Обновляем заголовки с учетом нового порядка столбцов
        ws.append(["Сеть", "Дата", "Время", "Сумма", "Валюта", "Комментарий", "Видимое", "", "Сеть", "Валюта", "Итого"])
        
        # Вставляем первую строку, потому что ее нужно обрабатывать по особенному
        ws.append(["=IF(NOT(ISBLANK(Промежуточный!A1)),Промежуточный!B1,\"\")",
                   "=IF(NOT(ISBLANK(Промежуточный!E1)),TEXT(Промежуточный!E1,\"DD.MM.YY\"),\"\")",
                   "=IF(NOT(ISBLANK(Промежуточный!E1)),TEXT(Промежуточный!E1,\"hh:mm\"),\"\")",
                   "=IF(NOT(ISBLANK(Промежуточный!D1)),Промежуточный!D1,\"\")",
                   "=IF(NOT(ISBLANK(Промежуточный!C1)),Промежуточный!C1,\"\")",
                   "=IF(NOT(ISBLANK(Промежуточный!G1)),Промежуточный!G1,\"\")",
                   "=IF(NOT(ISBLANK(Промежуточный!A1)),Промежуточный!B1,\"\")"
                   ])

        for col in "ABCDEF":
            cell = ws[f"{col}{6}"]
            cell.border = border

        for col in "DEF":
            cell = ws[f"{col}{6}"]
            cell.alignment = Alignment(horizontal="right")

        # Основные формулы для обработки данных
        for i in range(7, 6 + records_length):
            find_number = i - 5
            prev_number = i - 1
            ws[f"A{i}"] = ArrayFormula(f"A{i}:A{i}",
                f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!B{find_number}<>\"\",IF(NOT(IFERROR(MATCH("
                f"Промежуточный!B{find_number},$A$5:$A{prev_number},0),0)),"
                f"Промежуточный!B{find_number},\"\"),\"\"),\"\"),\"\")")
            ws[f"B{i}"] = f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!E{find_number}<>\"\"," \
                       f"IF(OR(TEXT(Промежуточный!E{find_number - 1},\"DD.MM.YY\")<>TEXT(" \
                       f"Промежуточный!E{find_number},\"DD.MM.YY\"),A{i}<>\"\"),TEXT(" \
                       f"Промежуточный!E{find_number},\"DD.MM.YY\"),\"\"),\"\"),\"\"),\"\")"
            ws[f"C{i}"] = f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!E{find_number}<>\"\"," \
                       f"TEXT(Промежуточный!E{find_number},\"hh:mm\"),\"\"),\"\"),\"\")"
            ws[f"D{i}"] = f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!D{find_number}<>\"\"," \
                       f"Промежуточный!D{find_number},\"\"),\"\"),\"\")"
            ws[f"E{i}"] = f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!C{find_number}<>\"\"," \
                       f"Промежуточный!C{find_number},\"\"),\"\"),\"\")"
            ws[f"F{i}"] = f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!G{find_number}<>\"\"," \
                       f"Промежуточный!G{find_number},\"\"),\"\"),\"\")"
            ws[f"G{i}"] = f"=IF(NOT(Промежуточный!I1),IFERROR(IF(Промежуточный!B{find_number}<>\"\"," \
                       f"Промежуточный!B{find_number},\"\"),\"\"),\"\")"

            # Применяем границы ко всем ячейкам в строке
            for col in "ABCDEF":
                cell = ws[f"{col}{i}"]
                cell.border = border

            # Применяем выравнивание по правому краю к сумме, валюте и комментарию
            for col in "DEF":
                cell = ws[f"{col}{i}"]
                cell.alignment = Alignment(horizontal="right")

        # Обновляем стили
        first_row = ws[1]
        for cell in first_row[:2]:
            cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb='fff2cc'))
            cell.border = border

        second_row = ws[2]
        for cell in second_row[:2]:
            cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb='fce4d6'))
            cell.border = border

        third_row = ws[3]
        for cell in third_row[:2]:
            cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb='e2efda'))
            cell.border = border

        header_row = ws[5]
        for cell in header_row[:7]:
            cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb='a9d08e'))
            cell.border = border
        for cell in header_row[8:11]:
            cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb='8ea9db'))
            cell.border = border

        max_currencies = sum([len(section_currencies[s]) for s in section_currencies]) + 5

        ws['L6'] = ArrayFormula(f"L6:M{max_currencies}",
                                f"=IFERROR(N6:O{max_currencies},\"\")")
        ws['N6'] = ArrayFormula(f"N6:O{max_currencies}",
                                f"=_xlfn.UNIQUE(Промежуточный!B1:C{records_length})")
        for row in ws.iter_rows(min_row=6, max_row=max_currencies, min_col=9, max_col=13):
            max_ind = records_length + 5
            row[0].value = f"=IF(L{row[0].row - 1}=L{row[0].row},\"\",L{row[0].row})"
            row[1].value = f"=IF(AND(M{row[1].row - 1}=M{row[1].row},I{row[1].row}=\"\"),\"\",M{row[1].row})"
            sumifs = f"SUMIFS($D$6:$D${max_ind},$E$6:$E${max_ind},$M{row[2].row},$G$6:$G${max_ind},$L{row[2].row})"
            row[2].value = f"=IF(AND(NOT(ISERROR($N{row[2].row})),NOT($J{row[2].row}=\"\")),{sumifs},\"\")"
            for cell in row[:3]:
                cell.border = border
            for cell in row[1:3]:
                cell.alignment = Alignment(horizontal="right")


